import openpyxl
import random

wb_obj = openpyxl.load_workbook("9input.xlsx")
sheet_obj = wb_obj.active
n = sheet_obj.max_row-1


def find_correlation(tid1, tid2):
    tid1_count = 0
    tid2_count = 0
    total_common_count = 0
   
    for j in range(2, 9):
        if (sheet_obj.cell(row=tid1+1, column=j).value) == "Y":
            tid1_count += 1
        if (sheet_obj.cell(row=tid2+1, column=j).value) == "Y":
            tid2_count += 1
        if ((sheet_obj.cell(row=tid1+1, column=j).value) == "Y") and ((sheet_obj.cell(row=tid2+1, column=j).value) == "Y"):
            total_common_count += 1
    if (tid1_count == 0 or tid2_count == 0):
            return 0

    return total_common_count/(tid1_count * tid2_count)


data = []
for i in range(1, n+1):
    for j in range(i+1, n+1):
        ans = find_correlation(i, j)
        if (ans == 0):
            verdict = "No relationship between entities"
        elif (ans < 0):
            verdict = "Negative correlation"
        elif (ans > 0):
            verdict = "Positive correlation"
        else:
            verdict = "Not defined"

        print("Correlation ratio " + str(i) + " & " +
            str(j) + " = " + str(ans) + " " + verdict + "\n")
        list = [str(i), str(j), str(ans), verdict]
        data.append(list)
# writing answer to file

workbook = openpyxl.Workbook()
sheet_obj = workbook.active

sheet_obj.cell(row=1, column=1).value = "item 1 with tid"
sheet_obj.cell(row=1, column=2).value = "item 2 with tid"
sheet_obj.cell(row=1, column=3).value = "Correlation coefficient"
sheet_obj.cell(row=1, column=4).value = "Type of correlation"

for i in range(0, len(data)):
    for j in range(0, len(data[i])):
        sheet_obj.cell(row=i+2, column=j+1).value = data[i][j]

workbook.save("Correlation_output.xlsx")